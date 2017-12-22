#!/usr/bin/env python
# -*- coding: UTF-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import sys
import time
import copy
import re
import json

reload(sys)
sys.setdefaultencoding('utf-8')


def getAllDir():
    """
    获取当前路径下的所有文件夹名字和创建时间
    :rtype:list 
    """
    folders = []
    for folder in os.listdir(sys.path[0]):
        if os.path.isdir(folder) and isNeedFolder(folder):
            file_info = []
            file_info.append(folder)
            file_info.append(get_dir_create_time(folder))
            folders.append(file_info)

    return folders

def isNeedFolder(folder):
    """
    
    :type folder: str
    :rtype: bool
    """
    if folder.endswith("众安换人".decode("utf-8")):
        return True
    return False

def get_dir_create_time(path):
    time_struct = time.localtime(os.path.getctime(path))
    return time.strftime('%Y-%m-%d', time_struct)



def getAllFile(dirs):
    """
    
    :type dirs: list
    :rtype: list
    """
    all_files = []
    for edir in dirs:
        dir_path = "{0}/{1}".format(sys.path[0], edir[0])
        files = []
        for file in os.listdir(dir_path):
            if file.endswith("xlsx"):
                files.append(file)

        edirs = copy.deepcopy(edir) #type: list
        edirs.append(files)

        all_files.append(edirs)
    return all_files

def getProtectNumberForFile(path):
    """
    
    :type path: str
    :rtype: str   保单号 
    """
    wb = load_workbook(path)
    ws = wb["增加被保险人".decode("utf-8")]
    cel = ws["H1"]
    #print "h1.cel:{0}".format(cel.value.encode("utf-8"))
    return re.findall(r"[a-zA-Z0-9]+", cel.value.encode("utf-8"))[0]


def makeExcelIndex(files):
    """
    
    :type files: list  [["17-8-13众安换人", "2017-12-14", ["xlsx","xlsx"]],...]
    :rtype: list   [["17-8-13众安换人", "2017-12-14", {"保单号":"xlsx","保单号":"xlsx"}],...]
    """
    rList = []
    for fileT in files:
        fileL = []
        fileL.append(fileT[0])
        fileL.append(fileT[1])
        if len(fileT[2]) > 0:
            fileD = {}
            for file in fileT[2]:
                fpath = "{0}/{1}/{2}".format(sys.path[0], fileT[0], file)
                protect_number = getProtectNumberForFile(fpath)
                fileD[protect_number] = file

            fileL.append(fileD)
        rList.append(fileL)

    return rList


def getSheetAllData(ws):
    """
    获取sheet 所有数据
    :type ws: `openpyxl.worksheet.Worksheet`
    :rtype: list
    """
    data = []

    col = 11 if ws.max_column > 11 else ws.max_column

    for row_iterate in ws.iter_rows(None, ws.min_row, ws.max_row, ws.min_column, col):
        every_row = []
        add_in = True
        for cel in row_iterate:
            every_row.append((cel.coordinate, cel.value))
            if cel.column == 'B' and cel.value == None and int(cel.row) > 4:
                add_in = False
                break


        if add_in:
            data.append(every_row)

    return data

def searchsheetCol(data, search_str):
    """
    获取查询的title行数据
    :type data: list
    :type search_str: str
    :rtype: tuple (row, col, 值)
    """
    for idx, every_row_data in enumerate(data):
        for num, value in enumerate(every_row_data):
            if value[1] and value[1] == search_str.decode("utf-8"):
                return idx, num, value

def getEveryRowsSearch(search_tuple, data):
    """
    获取每一行的查询用的key
    :type search_tuple: tuple(1, 4, ["E2", "起保日期"])
    :type data: list
    :rtype: list
    """

    all_of_result = []
    for search_num in range(1, len(data) - 1):
        idx = search_tuple[0] + search_num
        all_of_result.append((idx, data[idx][search_tuple[1]]))

    return all_of_result


def getDateString(dateStr):
    """
    匹配字符串中的时间字符串
    :type dateStr: str
    :rtype: str
    """
    dateList = []
    if dateStr:
        dateList = re.findall(r"\d{2,4}[-|.]\d+[-|.]\d+", dateStr)

    if len(dateList) > 0:
        return dateList[0]
    return ""

def compareDateEqual(date1, date2, isStart):

    """
    比较时间是否相等
    :type date1: str 在保日期
    :type date2: str 文件夹日期
    :type isStart: bool 是否启保
    :rtype: bool
    """
    if len(date1) == 0:
        return False
    if len(date2) == 0:
        return False

    d1 = date1.replace(".", "-")
    d2 = date2.replace(".", "-")
    datel1 = d1.split("-")
    datel2 = d2.split("-")

    #假设 数组count = 2 意味着没有年份,如果是没有月份或者没有日期 就是文件夹命名有问题
    num1 = len(datel1)
    num2 = len(datel2)
    date_compare_result = False
    if num1 == 2:
        datel1.insert(0, time.localtime()[0])
    if num2 == 2:
        datel2.insert(0, time.localtime()[0])


    # print "比较年份"
    y1 = datel1[0]
    y2 = datel2[0]
    year_compare_result = False
    if len(y1) == len(y2) and int(y1) == int(y2):
        year_compare_result = True
    elif len(y1) != len(y2):
        y1 = y1[-2:]
        y2 = y2[-2:]
        # print "y1:{0} y2:{0}".format(y1, y2)
        if int(y1) == int(y2):
            year_compare_result = True

    # print "比较月份"
    m1 = datel1[1]
    m2 = datel2[1]
    month_compare_result = False
    if int(m1) == int(m2):
        month_compare_result = True

    # print "比较日期"
    d1 = datel1[2]
    d2 = datel2[2]
    day_compare_result = False
    #在保日期 = 是文件夹日期 +1 终保日期 = 文件夹日期
    if isStart:
        d2 = int(d2) + 1

    if int(d1) == int(d2):
        day_compare_result = True

    if year_compare_result and month_compare_result and day_compare_result:
        date_compare_result = True

    return date_compare_result


def getFilePathWithStartProtectDate(all_dir, protect_date, basic_protect_number):
    """
    
    :type all_dir: list [["文件夹名字","文件夹创建时间",{"保单号":"文件名",...}],...]
    :type protect_date: str   起保时间
    :rtype: str  基础数据中的起保日期对应的文件夹中对应保单的文件路径
    """
    for edir in all_dir:
        d1 = getDateString(protect_date)
        d2 = getDateString(edir[0])
        if compareDateEqual(d1, d2, True):
            edict = edir[2] #type: dict
           # print json.dumps(edict, encoding="UTF-8", ensure_ascii=False)
            for key, value in edict.items():
                if key == basic_protect_number:
                    fpath = "{0}/{1}/{2}".format(sys.path[0], edir[0], value)
                    return fpath

    return None


def getFilePathWithEndProtectDate(all_dir, end_protect_date, basic_protect_number):
    """

    :type all_dir: list [["文件夹名字","文件夹创建时间",{"保单号":"文件名",...}],...]
    :type end_protect_date: str   终保时间
    :rtype: str  基础数据中的终保日期对应的文件夹中对应保单的文件路径
    """
    for edir in all_dir:
        d1 = getDateString(end_protect_date)
        d2 = getDateString(edir[0])
        if compareDateEqual(d1, d2, False):
            edict = edir[2]  # type: dict
            # print json.dumps(edict, encoding="UTF-8", ensure_ascii=False)
            for key, value in edict.items():
                if key == basic_protect_number:
                    fpath = "{0}/{1}/{2}".format(sys.path[0], edir[0], value)
                    return fpath

    return None


def getBasicExcelSheet():
    wb = load_workbook("基础数据表单.xlsx".decode("utf-8"))
    ws = wb["基础数据表单".decode("utf-8")]

    return ws



tdirs = getAllDir()
#print json.dumps(tdirs, encoding="UTF-8", ensure_ascii=False)
tfiles = getAllFile(tdirs)
#print json.dumps(tfiles, encoding="UTF-8", ensure_ascii=False)


mList = makeExcelIndex(tfiles)
#print json.dumps(mList, encoding="UTF-8", ensure_ascii=False)

bws = getBasicExcelSheet()
beList = getSheetAllData(bws)
#print json.dumps(beList, encoding="UTF-8", ensure_ascii=False)


#查询起保日期
sscList = searchsheetCol(beList, "起保日期".decode("utf-8"))
#print json.dumps(sscList, encoding="UTF-8", ensure_ascii=False)

ersList = getEveryRowsSearch(sscList,beList)
#print json.dumps(ersList, encoding="UTF-8", ensure_ascii=False)

#查询终保日期
sscList2 = searchsheetCol(beList, "退保终止日期".decode("utf-8"))
#print json.dumps(sscList2, encoding="UTF-8", ensure_ascii=False)

ersList2 = getEveryRowsSearch(sscList2,beList)
#print json.dumps(ersList2, encoding="UTF-8", ensure_ascii=False)

#获取姓名
sncList = searchsheetCol(beList, "姓名".decode("utf-8"))
esncList = getEveryRowsSearch(sncList, beList)
#print json.dumps(esncList, encoding="UTF-8", ensure_ascii=False)

#合并起保，终保
lastDates = []
for idx in range(0, len(ersList), 1):
    l1 = ersList[idx]
    l2 = ersList2[idx]
    l3 = esncList[idx]
    edates = []
    edates.append(l1[0])
    edates.append(l1[1])
    edates.append(l2[1])
    edates.append(l3[1])
    lastDates.append(edates)


#print json.dumps(lastDates, encoding="UTF-8", ensure_ascii=False)

#通过日期查询文件名字
#基础保单号
basicPrdNumber = re.findall(r"[a-zA-Z0-9]+", beList[0][0][1])[0]
#print "基础数据的保单号: {0}".format(basicPrdNumber)

errorData = []

for epro_date in lastDates:
    del_path = getFilePathWithEndProtectDate(mList, epro_date[2][1], basicPrdNumber)
    add_path = getFilePathWithStartProtectDate(mList, epro_date[1][1], basicPrdNumber)
    #print "根据起保日期查询到的文件路径： %s" %fpath
    if del_path:
        #print json.dumps(beList[epro_date[0]], encoding="UTF-8", ensure_ascii=False)
        del_data_is_error = False
        ecwb = load_workbook(del_path)
        ec_delws = ecwb["减少被保险人".decode("utf-8")]
        ec_delList = getSheetAllData(ec_delws)
        for test in ec_delList[1:]:
            #print json.dumps(test, encoding="UTF-8", ensure_ascii=False)
            del_name = test[1][1] #type: str

            if del_name and del_name == epro_date[3][1]:
                sdate1 = test[5][1]
                sdate2 = getDateString(epro_date[1][1])
                if sdate1:
                    sdate1 = re.findall(r"\d{2,4}\d+\d+", sdate1)[0]
                    sdate2 = sdate2.replace("-", "")
                    if sdate1 == sdate2:
                        del_data_is_error = False
                        break
                    else:
                        del_data_is_error = True
            else:
                del_data_is_error = True

        if del_data_is_error:
             # print "错误数据1:"
             # print json.dumps(beList[epro_date[0]], encoding="UTF-8", ensure_ascii=False)
             errorData.append(beList[epro_date[0]])


    elif add_path:
        add_data_is_error = False
        ecwb = load_workbook(add_path)
        ec_addws = ecwb["增加被保险人".decode("utf-8")]
        ec_addList = getSheetAllData(ec_addws)

        for test1 in ec_addList[2:]:
            add_name = test1[1][1] #type: str

            if add_name and add_name == epro_date[3][1]:
                sadate1 = test1[8][1]
                sadate2 = getDateString(epro_date[1][1])
                if sadate1:
                    sadate1 = re.findall(r"\d{2,4}\d+\d+", sadate1)[0]
                    sadate2 = sadate2.replace("-", "")
                    if sadate1 == sadate2:
                        add_data_is_error = False
                        break
                    else:
                        add_data_is_error = True
            else:
                add_data_is_error = True

        if add_data_is_error:
            # print "错误数据2:"
            # print json.dumps(beList[epro_date[0]], encoding="UTF-8", ensure_ascii=False)
            errorData.append(beList[epro_date[0]])

    else:
        # print "错误数据3:"
        # print json.dumps(beList[epro_date[0]], encoding="UTF-8", ensure_ascii=False)
        errorData.append(beList[epro_date[0]])




errorData.insert(0, beList[0])
errorData.insert(1, beList[1])


errorws = Workbook()
errorws.create_sheet("错误数据".decode("utf-8"), index=0)
errorsheet = errorws.active

for error_row in errorData:
    for error_col in error_row:
        indexName = error_col[0]
        indexValue = error_col[1]
        if indexValue:
            errorsheet[indexName] = indexValue

error_path = "{0}/{1}".format(sys.path[0], "保单错误数据.xlsx".decode("utf-8"))
errorws.save(error_path)






