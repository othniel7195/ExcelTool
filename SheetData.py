#!/usr/bin/env python
# -*- coding: UTF-8 -*-

from openpyxl import Workbook

errorws = Workbook(write_only=True)
errorws.create_sheet(u"错误数据", index=0)
print errorws.sheetnames

errorsheet = errorws.active
print errorsheet

errorsheet["A1"] = "1111"


































